# -*- coding: utf-8 -*-
"""
AI Maintenance Tracker — Final Single-File App
- Arabic RTL light UI (colorful)
- DB migration safe (SQLite)
- Locations/Teams/Assets/Visits/Alerts
- Oil interval (hours) + daily operating hours per site
- Auto compute next oil due date (equation)
- Predictions page + Excel export
- Monthly / period Excel export with header
- Capacity-aware fortnight planner
- KPIs + simple PWA cache
"""

from __future__ import annotations
import os, io, math
from datetime import date, datetime, timedelta
from typing import Optional, List

from fastapi import FastAPI, Depends, HTTPException, Header, Query
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse, PlainTextResponse
from fastapi.middleware.cors import CORSMiddleware

from pydantic import BaseModel
from sqlmodel import SQLModel, Field, Session, select, create_engine
from sqlalchemy import text

from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

APP_NAME = "AI Maintenance Tracker"
API_KEY = os.getenv("API_KEY")  # Optional: require X-API-Key for POST/PUT if set
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///ai_tracker.db")
connect_args = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
engine = create_engine(DATABASE_URL, echo=False, connect_args=connect_args)
from fastapi import FastAPI
app = FastAPI(title=APP_NAME)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

# ------------------------- Auth -------------------------
def require_api_key(x_api_key: Optional[str] = Header(default=None)):
    if API_KEY and x_api_key != API_KEY:
        raise HTTPException(401, "Invalid or missing API key")
    return True

# ------------------------- Models -------------------------
class Team(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    name: str

class Location(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    name: str
    region: str
    site_owner: Optional[str] = None

class Asset(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    location_id: int = Field(foreign_key="location.id")

    generator_model: Optional[str] = None
    generator_serial: Optional[str] = None
    secondary_power: str = "مولد + كهرباء عمومية"

    hourmeter: int = 0

    # last oil change snapshot
    last_oil_change_date: Optional[date] = None
    last_oil_change_hours: Optional[int] = None

    # manufacturer recommendation — interval by HOURS
    oil_interval_hours: int = 250

    # average running hours per day (used for prediction)
    daily_operating_hours: int = 6

    # auto-computed: predicted next oil change date
    next_oil_due_date: Optional[date] = None

class Visit(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    date: datetime = Field(default_factory=datetime.utcnow, index=True)
    location_id: int = Field(foreign_key="location.id")
    team_id: int = Field(foreign_key="team.id")
    visit_type: str  # periodic|emergency|inspection|supply|other
    summary: str = ""
    hourmeter: int = 0
    grid_meter: int = 0
    delta_hours: int = 0
    i_l1: float = 0
    i_l2: float = 0
    i_l3: float = 0
    performed_oil_change: bool = False
    executor_name: Optional[str] = None
    driver_name: Optional[str] = None
    site_owner: Optional[str] = None
    emergency_source: Optional[str] = None
    emergency_alarm: Optional[str] = None
    emergency_class: Optional[str] = None
    notes: Optional[str] = None

class ConsumableUsed(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    visit_id: int = Field(foreign_key="visit.id")
    name: str
    qty: float = 0

class Alert(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    location_id: int = Field(foreign_key="location.id")
    kind: str  # due_oil|current_unbalance|current_anomaly|counter_inconsistent
    level: str = "info"  # info|warn|critical
    message: str
    created_at: datetime = Field(default_factory=datetime.utcnow)
    is_open: bool = True

class InventoryItem(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    name: str
    unit: str = ""
    stock: float = 0
    min_stock: float = 0

def get_session():
    with Session(engine) as s:
        yield s

# ------------------------- DB Migration (safe) -------------------------
def _col_missing(conn, table, col):
    res = conn.execute(text(f"PRAGMA table_info({table});")).fetchall()
    names = {r[1] for r in res}
    return col not in names

def migrate_db():
    with engine.begin() as conn:
        # Location
        if _col_missing(conn, "location", "site_owner"):
            conn.execute(text("ALTER TABLE location ADD COLUMN site_owner TEXT"))
        # Asset
        if _col_missing(conn, "asset", "generator_model"):
            conn.execute(text("ALTER TABLE asset ADD COLUMN generator_model TEXT"))
        if _col_missing(conn, "asset", "generator_serial"):
            conn.execute(text("ALTER TABLE asset ADD COLUMN generator_serial TEXT"))
        if _col_missing(conn, "asset", "secondary_power"):
            conn.execute(text("ALTER TABLE asset ADD COLUMN secondary_power TEXT DEFAULT 'مولد + كهرباء عمومية'"))
        if _col_missing(conn, "asset", "hourmeter"):
            conn.execute(text("ALTER TABLE asset ADD COLUMN hourmeter INTEGER DEFAULT 0"))
        if _col_missing(conn, "asset", "last_oil_change_date"):
            conn.execute(text("ALTER TABLE asset ADD COLUMN last_oil_change_date DATE"))
        if _col_missing(conn, "asset", "last_oil_change_hours"):
            conn.execute(text("ALTER TABLE asset ADD COLUMN last_oil_change_hours INTEGER"))
        if _col_missing(conn, "asset", "oil_interval_hours"):
            conn.execute(text("ALTER TABLE asset ADD COLUMN oil_interval_hours INTEGER DEFAULT 250"))
        if _col_missing(conn, "asset", "daily_operating_hours"):
            conn.execute(text("ALTER TABLE asset ADD COLUMN daily_operating_hours INTEGER DEFAULT 6"))
        if _col_missing(conn, "asset", "next_oil_due_date"):
            conn.execute(text("ALTER TABLE asset ADD COLUMN next_oil_due_date DATE"))

def init_db():
    # seed minimal data if empty
    with Session(engine) as s:
        if not s.exec(select(Team)).first():
            s.add_all([Team(name="الفريق الأول"), Team(name="الفريق الثاني"), Team(name="الفريق الثالث")])
            s.commit()
        if not s.exec(select(Location)).first():
            loc = Location(name="النهضة 1", region="الأمانة", site_owner="الجهة أ")
            s.add(loc); s.commit(); s.refresh(loc)
            s.add(Asset(location_id=loc.id, generator_model="Gen-50kVA",
                        generator_serial="SN-0001",
                        secondary_power="مولد + كهرباء عمومية",
                        hourmeter=3000, last_oil_change_hours=3000,
                        last_oil_change_date=date.today()-timedelta(days=15),
                        oil_interval_hours=250, daily_operating_hours=6))
            s.add_all([
                InventoryItem(name="زيت محرك", unit="لتر", stock=400, min_stock=80),
                InventoryItem(name="فلتر زيت", unit="pcs", stock=60, min_stock=20),
                InventoryItem(name="فلتر ديزل", unit="pcs", stock=60, min_stock=20),
                InventoryItem(name="فلتر هواء", unit="pcs", stock=40, min_stock=10),
                InventoryItem(name="سير مولد", unit="pcs", stock=12, min_stock=5),
            ])
            s.commit()

# ------------------------- Logic & Helpers -------------------------
def compute_next_oil_due(asset: Asset) -> Optional[date]:
    """
    days_to_due = ceil( (interval_hours - (hourmeter - last_oil_change_hours)) / daily_operating_hours )
    """
    if asset.last_oil_change_hours is None or asset.daily_operating_hours is None:
        return None
    used_since = max(0, (asset.hourmeter or 0) - (asset.last_oil_change_hours or 0))
    remaining_hours = max(0, (asset.oil_interval_hours or 250) - used_since)
    doh = max(1, asset.daily_operating_hours or 1)
    days_to_due = math.ceil(remaining_hours / doh)
    return date.today() + timedelta(days=days_to_due)

def due_oil(asset: Asset, current_hourmeter: Optional[int] = None) -> bool:
    hm = asset.hourmeter if current_hourmeter is None else current_hourmeter
    if asset.last_oil_change_hours is not None:
        return (hm - (asset.last_oil_change_hours or 0)) >= (asset.oil_interval_hours or 250)
    if asset.last_oil_change_date and asset.daily_operating_hours:
        est_hours = (date.today() - asset.last_oil_change_date).days * max(1, asset.daily_operating_hours)
        return est_hours >= (asset.oil_interval_hours or 250)
    return False

def check_current_unbalance(i1: float, i2: float, i3: float):
    vals = [abs(i1), abs(i2), abs(i3)]
    avg = sum(vals)/3 if any(vals) else 0
    if avg == 0:
        return None
    if min(vals) == 0 and max(vals) > 0:
        return ("critical", "تيار صفري في فاز مع وجود حمل")
    unb = (max(vals) - min(vals)) / avg
    if unb >= 0.2:
        return ("warn", f"عدم توازن ≥ {round(unb*100)}%")
    return None

def robust_current_anomaly(session: Session, location_id: int, triplet):
    rows = session.exec(
        select(Visit).where(Visit.location_id == location_id)
                     .where(Visit.visit_type == "periodic")
                     .order_by(Visit.date.desc()).limit(10)
    ).all()
    if len(rows) < 5:
        return None
    import statistics as st
    xs = [[v.i_l1 for v in rows], [v.i_l2 for v in rows], [v.i_l3 for v in rows]]
    new = list(triplet)
    for k in range(3):
        series = xs[k] + [new[k]]
        med = st.median(series)
        mad = st.median([abs(z-med) for z in series]) or 1.0
        score = abs(series[-1]-med)/(1.4826*mad)
        if score >= 4.0:
            return ("warn", "انحراف تيارات غير اعتيادي (تحليل تاريخي)")
    return None

def default_summary(vtype: str, name: str) -> str:
    m = {
        "periodic": "تم عمل صيانة للمولد وتجهيزات القوى والتكييف — " + name,
        "emergency": "تم التعامل مع بلاغ طارئ — " + name,
        "inspection": "تمت صيانة تفقدية — " + name,
        "supply": "تم توريد/استلام مواد — " + name,
        "other": "تم تنفيذ مهام أخرى — " + name,
    }
    return m.get(vtype, "زيارة — " + name)

# ------------------------- Schemas -------------------------
class TeamCreate(BaseModel): name: str
class TeamUpdate(BaseModel): name: Optional[str] = None

class LocationCreate(BaseModel):
    name: str; region: str; site_owner: Optional[str] = None
class LocationUpdate(BaseModel):
    name: Optional[str]=None; region: Optional[str]=None; site_owner: Optional[str]=None

class AssetUpdate(BaseModel):
    generator_model: Optional[str] = None
    generator_serial: Optional[str] = None
    secondary_power: Optional[str] = None
    hourmeter: Optional[int] = None
    last_oil_change_date: Optional[date] = None
    last_oil_change_hours: Optional[int] = None
    oil_interval_hours: Optional[int] = None
    daily_operating_hours: Optional[int] = None

class VisitCreate(BaseModel):
    location_id: int; team_id: int; visit_type: str
    hourmeter: int; grid_meter: int = 0
    i_l1: float = 0; i_l2: float = 0; i_l3: float = 0
    performed_oil_change: bool = False
    date: Optional[datetime] = None
    summary: Optional[str]=None; executor_name: Optional[str]=None
    driver_name: Optional[str]=None; site_owner: Optional[str]=None
    emergency_source: Optional[str]=None; emergency_alarm: Optional[str]=None
    emergency_class: Optional[str]=None; notes: Optional[str]=None
    oil_liters: float = 0; oil_filter: int = 0; diesel_filter: int = 0; air_filter: int = 0; belt_qty: int = 0

# ------------------------- CRUD -------------------------
@app.get("/health")
def health(): return {"ok": True, "app": APP_NAME, "time": datetime.utcnow().isoformat()}

@app.get("/teams")
def list_teams(session: Session = Depends(get_session)):
    return session.exec(select(Team).order_by(Team.name)).all()

@app.post("/teams")
def create_team(payload: TeamCreate, session: Session = Depends(get_session)):
    t = Team(**payload.dict()); session.add(t); session.commit(); session.refresh(t); return t

@app.put("/teams/{team_id}")
def update_team(team_id: int, payload: TeamUpdate, session: Session = Depends(get_session)):
    t = session.get(Team, team_id)
    if not t: raise HTTPException(404, "Team not found")
    for k,v in payload.dict(exclude_none=True).items(): setattr(t,k,v)
    session.add(t); session.commit(); session.refresh(t); return t

@app.get("/locations")
def list_locations(q: Optional[str] = None, limit: int = 1000, session: Session = Depends(get_session)):
    stmt = select(Location)
    if q:
        ql = f"%{q}%"
        stmt = stmt.where((Location.name.like(ql)) | (Location.region.like(ql)))
    stmt = stmt.order_by(Location.region, Location.name)
    rows = session.exec(stmt).all()
    return rows[:max(1, min(limit, 5000))]

@app.post("/locations")
def create_location(payload: LocationCreate, session: Session = Depends(get_session)):
    # تحقق من المدخلات
    name = (payload.name or "").strip()
    region = (payload.region or "").strip()
    if not name or not region:
        raise HTTPException(status_code=400, detail="يجب إدخال اسم الموقع والمنطقة")

    try:
        # إنشاء الموقع
        loc = Location(name=name, region=region, site_owner=(payload.site_owner or "").strip() or None)
        session.add(loc); session.commit(); session.refresh(loc)
        # إنشاء أصل افتراضي مرتبط بالموقع
        asset = Asset(location_id=loc.id)
        session.add(asset); session.commit()
        return {"ok": True, "id": loc.id, "name": loc.name}
    except Exception as e:
        # إرجاع سبب واضح (مهم أثناء التطوير)
        session.rollback()
        raise HTTPException(status_code=500, detail=f"خطأ قاعدة البيانات: {e}")


@app.put("/locations/{loc_id}")
def update_location(loc_id: int, payload: LocationUpdate, session: Session = Depends(get_session)):
    loc = session.get(Location, loc_id)
    if not loc: raise HTTPException(404, "Location not found")
    for k,v in payload.dict(exclude_none=True).items(): setattr(loc,k,v)
    session.add(loc); session.commit(); session.refresh(loc); return loc

@app.get("/assets/by_location")
def get_asset_by_location(location_id: int, session: Session = Depends(get_session)):
    a = session.exec(select(Asset).where(Asset.location_id == location_id)).first()
    if not a: raise HTTPException(404, "Asset not found")
    return a

@app.put("/assets/{asset_id}")
def update_asset(asset_id: int, payload: AssetUpdate, session: Session = Depends(get_session)):
    a = session.get(Asset, asset_id)
    if not a: raise HTTPException(404, "Asset not found")
    for k,v in payload.dict(exclude_none=True).items(): setattr(a,k,v)
    a.next_oil_due_date = compute_next_oil_due(a)  # recompute when inputs change
    session.add(a); session.commit(); session.refresh(a); return a

@app.post("/visits")
def create_visit(payload: VisitCreate, session: Session = Depends(get_session)):
    loc = session.get(Location, payload.location_id)
    team = session.get(Team, payload.team_id)
    if not loc or not team: raise HTTPException(400, "Invalid location/team")

    last = session.exec(select(Visit).where(Visit.location_id==payload.location_id).order_by(Visit.date.desc())).first()
    prev_h = last.hourmeter if last else 0
    delta = payload.hourmeter - prev_h
    if delta < 0 or delta > 240:
        session.add(Alert(location_id=payload.location_id, kind="counter_inconsistent", level="warn",
                          message="قراءة عداد غير منطقية"))
        delta = max(delta, 0)

    v = Visit(
        date = payload.date or datetime.utcnow(),
        location_id=payload.location_id, team_id=payload.team_id, visit_type=payload.visit_type,
        hourmeter=payload.hourmeter, grid_meter=payload.grid_meter, delta_hours=delta,
        i_l1=payload.i_l1, i_l2=payload.i_l2, i_l3=payload.i_l3,
        performed_oil_change=payload.performed_oil_change,
        summary=payload.summary or default_summary(payload.visit_type, loc.name),
        executor_name=payload.executor_name, driver_name=payload.driver_name,
        site_owner=payload.site_owner, emergency_source=payload.emergency_source,
        emergency_alarm=payload.emergency_alarm, emergency_class=payload.emergency_class,
        notes=payload.notes
    )
    session.add(v); session.commit(); session.refresh(v)

    # consumables
    def add(name, qty):
        if qty and float(qty)!=0: session.add(ConsumableUsed(visit_id=v.id, name=name, qty=float(qty)))
    add("الزيت(لتر)", payload.oil_liters)
    add("فلتر الزيت", payload.oil_filter)
    add("فلتر الديزل", payload.diesel_filter)
    add("فلتر الهواء", payload.air_filter)
    add("سير مولد", payload.belt_qty)
    session.commit()

    a = session.exec(select(Asset).where(Asset.location_id==payload.location_id)).first()
    if a:
        a.hourmeter = payload.hourmeter
        if payload.performed_oil_change:
            a.last_oil_change_date = date.today()
            a.last_oil_change_hours = payload.hourmeter
        a.next_oil_due_date = compute_next_oil_due(a)
        session.add(a); session.commit()

    ub = check_current_unbalance(payload.i_l1, payload.i_l2, payload.i_l3)
    if ub: session.add(Alert(location_id=payload.location_id, kind="current_unbalance", level=ub[0], message=ub[1]))
    an = robust_current_anomaly(session, payload.location_id, (payload.i_l1, payload.i_l2, payload.i_l3))
    if an: session.add(Alert(location_id=payload.location_id, kind="current_anomaly", level=an[0], message=an[1]))
    if a and due_oil(a, payload.hourmeter):
        session.add(Alert(location_id=payload.location_id, kind="due_oil", level="info", message="زيت/فلاتر مستحقة"))
    session.commit()
    return {"ok": True, "visit_id": v.id}

@app.get("/alerts")
def list_alerts(region: Optional[str] = None, only_open: bool = True, session: Session = Depends(get_session)):
    q = select(Alert)
    if only_open: q = q.where(Alert.is_open == True)
    alerts = session.exec(q.order_by(Alert.created_at.desc())).all()
    if region:
        loc_ids = [l.id for l in session.exec(select(Location).where(Location.region==region)).all()]
        alerts = [a for a in alerts if a.location_id in loc_ids]
    return alerts

@app.post("/alerts/{alert_id}/close")
def close_alert(alert_id: int, session: Session = Depends(get_session)):
    a = session.get(Alert, alert_id)
    if not a: raise HTTPException(404, "Alert not found")
    a.is_open = False; session.add(a); session.commit()
    return {"ok": True}

# ------------------------- Planner & KPIs -------------------------
def _last_visit(session: Session, location_id: int):
    return session.exec(select(Visit).where(Visit.location_id==location_id).order_by(Visit.date.desc())).first()
def _count_emerg(session: Session, location_id: int, days=30)->int:
    since = datetime.utcnow() - timedelta(days=days)
    q = select(Visit).where(Visit.location_id==location_id).where(Visit.visit_type=="emergency").where(Visit.date>=since)
    return len(session.exec(q).all())
def _has_open(session: Session, location_id: int, kind: str)->bool:
    q = select(Alert).where(Alert.location_id==location_id).where(Alert.kind==kind).where(Alert.is_open==True)
    return session.exec(q).first() is not None

def compute_location_health(session: Session, loc: Location)->dict:
    asset = session.exec(select(Asset).where(Asset.location_id==loc.id)).first()
    last = _last_visit(session, loc.id)
    days_since = None if not last else (date.today() - last.date.date()).days
    emerg30 = _count_emerg(session, loc.id, 30)
    f_days = 0 if days_since is None else min(days_since/20.0, 2.0)
    f_emerg = min(emerg30/3.0, 2.0)
    f_unb = 1.5 if _has_open(session, loc.id, "current_unbalance") else 0.0
    f_due = 1.0 if _has_open(session, loc.id, "due_oil") else 0.0
    f_an  = 1.2 if _has_open(session, loc.id, "current_anomaly") else 0.0
    linear = (2.0*f_days)+(2.0*f_emerg)+(2.5*f_unb)+(1.5*f_due)+(1.4*f_an)
    prob = 1/(1+math.exp(-linear)); score = int(round(prob*100))
    rec = 0 if score>=80 else (1 if score>=65 else (3 if score>=50 else 7))
    reasons = ["لا توجد زيارات سابقة" if days_since is None else f"آخر زيارة منذ {days_since} يوم"]
    if emerg30: reasons.append(f"{emerg30} طارئ (30ي)")
    if f_unb: reasons.append("عدم توازن التيار")
    if f_due: reasons.append("زيت/فلاتر مستحقة")
    if f_an: reasons.append("انحراف تيارات (تاريخي)")
    return {"location_id":loc.id,"location":loc.name,"region":loc.region,"score":score,
            "recommend_in_days":rec,"last_visit_at":None if not last else last.date.isoformat(),
            "recent_emergencies_30d":emerg30,"has_unbalance":bool(f_unb),"due_oil":bool(f_due),"has_anomaly":bool(f_an),
            "reasons":reasons}

@app.get("/kpi/overview")
def kpi_overview(session: Session = Depends(get_session)):
    total = session.exec(select(Location)).all()
    open_alerts = session.exec(select(Alert).where(Alert.is_open==True)).all()
    emerg_30 = session.exec(select(Visit).where(Visit.visit_type=="emergency")
                            .where(Visit.date>=datetime.utcnow()-timedelta(days=30))).all()
    overdue=0
    for l in total:
        last=_last_visit(session,l.id)
        if last and last.visit_type in ("periodic","inspection"):
            if (date.today()-last.date.date()).days>20: overdue+=1
        elif not last: overdue+=1
    return {"locations":len(total),"alerts_open":len(open_alerts),
            "emergency_30d":len(emerg_30),"overdue_periodic":overdue}

@app.get("/planner/today")
def plan_today(limit: int = 100, region: Optional[str] = None, session: Session = Depends(get_session)):
    q = select(Location); 
    if region: q = q.where(Location.region==region)
    items = [compute_location_health(session, l) for l in session.exec(q).all()]
    items = [i for i in items if i["recommend_in_days"] <= 1]
    items.sort(key=lambda x: (-x["score"], x["recommend_in_days"], x["region"], x["location"]))
    return items[:limit]

@app.get("/planner/fortnight")
def plan_fortnight(max_days: int = 14, team_capacity: int = 6, session: Session = Depends(get_session)):
    teams = session.exec(select(Team).order_by(Team.name)).all()
    locs = session.exec(select(Location)).all()
    items = [compute_location_health(session, l) for l in locs]
    items.sort(key=lambda x: -x["score"])
    start = date.today()
    days = [start + timedelta(days=i) for i in range(max_days)]
    schedule = {str(d): {t.id: [] for t in teams} for d in days}
    unassigned = []
    for it in items:
        first = 0 if it["recommend_in_days"]==0 else (1 if it["recommend_in_days"]==1 else (3 if it["recommend_in_days"]==3 else 7))
        placed = False
        for di in range(first, len(days)):
            best, load = None, 10**9
            for t in teams:
                L = len(schedule[str(days[di])][t.id])
                if L < team_capacity and L < load:
                    load = L; best = t.id
            if best is not None:
                schedule[str(days[di])][best].append(it); placed = True; break
        if not placed: unassigned.append(it)
    return {"start":str(start),"days":[str(d) for d in days],
            "teams":[{"id":t.id,"name":t.name} for t in teams],
            "team_capacity":team_capacity,"schedule":schedule,"unassigned":unassigned}

# ------------------------- Reports -------------------------
AR_HEADERS = ["م","التاريخ","المنطقة","الموقع","نوع العمل","العمل المنجز (ملخص فقط)",
              "عداد الساعات","فارق القراءه","الزيت(لتر)","فلتر الزيت","فلتر الديزل","فلتر الهواء","سير مولد",
              "التيار في الفاز الأول","التيار في الفاز الثاني","التيار في الفاز الثالث",
              "اسم القطعه","الكميه","الإنذار","مصدر البلاغ","تصنيف المشكلة","تبعية الموقع","المنفذ للعمل","السائق","ملاحظات"]

def _cons_map(session, vid:int):
    items=session.exec(select(ConsumableUsed).where(ConsumableUsed.visit_id==vid)).all()
    cols={"الزيت(لتر)":0,"فلتر الزيت":0,"فلتر الديزل":0,"فلتر الهواء":0,"سير مولد":0}
    others=[]
    for it in items:
        if it.name in cols: cols[it.name]=it.qty
        else: others.append(f"{it.name}×{it.qty}")
    spare="، ".join(others) if others else ""
    qty="" if not others else "-"
    return cols, spare, qty

def _date_parse(s: str) -> datetime:
    return datetime.strptime(s, "%Y-%m-%d")

@app.get("/reports/export")
def export_by_period(
    start: str = Query(..., description="YYYY-MM-DD"),
    end: str   = Query(..., description="YYYY-MM-DD (exclusive)"),
    region: Optional[str] = None,
    session: Session = Depends(get_session)
):
    start_dt = _date_parse(start); end_dt = _date_parse(end)
    q = (select(Visit, Location)
         .where(Visit.location_id == Location.id)
         .where(Visit.date >= start_dt)
         .where(Visit.date <  end_dt))
    if region: q = q.where(Location.region==region)
    rows = session.exec(q.order_by(Visit.date, Location.region, Location.name)).all()

    wb=Workbook(); ws=wb.active; ws.title="تقرير الفترة"; ws.sheet_view.rightToLeft=True
    title = "التقرير للفترة: من " + str(start_dt.date()) + " إلى " + str((end_dt - timedelta(days=1)).date())
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(AR_HEADERS))
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, size=14, color="1d3557")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.append([])

    ws.append(AR_HEADERS); thin=Side(style="thin", color="b3c1d1")
    for c in ws[3]:
        c.font=Font(bold=True, color="1d3557")
        c.fill=PatternFill(start_color="dff2ff", end_color="dff2ff", fill_type="solid")
        c.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border=Border(left=thin,right=thin,top=thin,bottom=thin)

    map_type={"periodic":"صيانة دورية","emergency":"صيانة طارئة","inspection":"صيانة تفقدية","supply":"توريد مواد"}
    for idx,(v,l) in enumerate(rows,1):
        cols,spare,qty=_cons_map(session,v.id)
        ws.append([idx,v.date.date().isoformat(),l.region,l.name,map_type.get(v.visit_type,"مهام أخرى"),
                   v.summary or "",v.hourmeter,v.delta_hours,
                   cols["الزيت(لتر)"],cols["فلتر الزيت"],cols["فلتر الديزل"],cols["فلتر الهواء"],cols["سير مولد"],
                   v.i_l1,v.i_l2,v.i_l3,spare,qty,
                   v.emergency_alarm or "", v.emergency_source or "", v.emergency_class or "",
                   v.site_owner or l.site_owner or "", v.executor_name or "", v.driver_name or "", v.notes or ""])

    widths=[5,12,12,18,14,28,12,12,10,10,10,10,10,12,12,12,18,10,14,16,16,16,14,14,20]
    for i,w in enumerate(widths,1):
        col = chr(64+i) if i<=26 else "A" + str(i-26)
        ws.column_dimensions[col].width=w
    for row in ws.iter_rows(min_row=4,max_row=ws.max_row,min_col=1,max_col=len(AR_HEADERS)):
        for c in row:
            c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            c.border=Border(left=thin,right=thin,top=thin,bottom=thin)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": 'attachment; filename="Report-{}-to-{}.xlsx"'.format(start, end)})

@app.get("/reports/monthly")
def export_monthly(month: str, region: Optional[str] = None, session: Session = Depends(get_session)):
    y,m = [int(x) for x in month.split("-")]
    start = datetime(y,m,1)
    end = (start+timedelta(days=32)).replace(day=1)
    return export_by_period(start=start.date().isoformat(), end=end.date().isoformat(), region=region, session=session)

# ------------------------- Predictions (API + Page + Export) -------------------------
@app.get("/predictions/next_month")
def predictions_next_month(session: Session = Depends(get_session)):
    today = date.today(); horizon = today + timedelta(days=30)
    data = []
    locs = session.exec(select(Location)).all()
    for l in locs:
        a = session.exec(select(Asset).where(Asset.location_id == l.id)).first()
        if not a: continue
        due = a.next_oil_due_date or compute_next_oil_due(a)
        in_window = (due is not None) and (today <= due <= horizon)
        days_left = None if not due else (due - today).days
        data.append({
            "location_id": l.id, "location": l.name, "region": l.region,
            "hourmeter": a.hourmeter, "interval_hours": a.oil_interval_hours,
            "daily_hours": a.daily_operating_hours, "last_oil_hours": a.last_oil_change_hours,
            "last_oil_date": None if not a.last_oil_change_date else a.last_oil_change_date.isoformat(),
            "next_due_date": None if not due else due.isoformat(), "within_30_days": bool(in_window),
            "days_left": days_left,
        })
    data.sort(key=lambda x: (not x["within_30_days"], 9999 if x["days_left"] is None else x["days_left"]))
    return data

@app.get("/predictions/export")
def predictions_export(session: Session = Depends(get_session)):
    today = date.today()
    locs = session.exec(select(Location)).all()
    data = []
    for l in locs:
        a = session.exec(select(Asset).where(Asset.location_id == l.id)).first()
        if not a: continue
        due = a.next_oil_due_date or compute_next_oil_due(a)
        days_left = None if not due else (due - today).days
        data.append([l.region, l.name, a.hourmeter, a.oil_interval_hours, a.daily_operating_hours,
                     a.last_oil_change_hours or "", a.last_oil_change_date.isoformat() if a.last_oil_change_date else "",
                     due.isoformat() if due else "", "" if days_left is None else days_left])

    wb = Workbook(); ws = wb.active; ws.title = "Predictions"
    ws.sheet_view.rightToLeft = True
    headers = ["المنطقة","الموقع","عداد الساعات","فاصل الشركة (س)","تشغيل يومي (س)",
               "آخر تبديل زيت (س)","آخر تبديل زيت (ت)","الموعد القادم","أيام متبقية"]
    ws.append(headers)
    for row in data: ws.append(row)
    for i,w in enumerate([14,20,14,16,16,18,18,18,14], start=1):
        col = chr(64+i) if i<=26 else "A"+str(i-26)
        ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="PredictionsNextMonth.xlsx"'}
    )

# ------------------------- Styles -------------------------
BASE_STYLE = """
:root{
  --bg:#f6f8fc; --card:#ffffff; --ink:#1d3557; --sub:#5b6b8b; --line:#e3e8f2;
  --accent:#4f46e5; --accent-2:#10b981; --accent-3:#06b6d4; --warn:#f59e0b; --crit:#ef4444
}
*{box-sizing:border-box}
body{margin:0;background:var(--bg);color:var(--ink);font-family:Segoe UI, Tahoma, Arial}
a{color:var(--accent)}
.wrap{max-width:1280px;margin:18px auto;padding:0 12px}
.row{display:flex;gap:12px;flex-wrap:wrap}
.col{flex:1 1 260px;min-width:260px}
.card{background:var(--card);border:1px solid var(--line);border-radius:16px;padding:14px;box-shadow:0 6px 18px rgba(0,0,0,.05);margin-bottom:14px}
label{display:block;font-weight:600;margin:6px 0;color:#334155}
input,select,textarea{width:100%;padding:12px;border-radius:12px;border:1px solid var(--line);background:#fbfdff;color:var(--ink)}
input:focus,select:focus,textarea:focus{outline:2px solid #c7d2fe}
textarea{min-height:76px}
.btn{border:0;border-radius:12px;padding:10px 14px;background:var(--accent);color:#fff;cursor:pointer}
.btn.secondary{background:var(--accent-3)}
.btn.ok{background:var(--accent-2)}
.btn.warn{background:var(--warn)}
.btn.ghost{background:#fff;border:1px solid var(--line);color:#0f172a}
.inline-actions{display:flex;gap:8px;align-items:flex-end}
.minibtn{padding:9px 10px;border-radius:12px;border:1px solid var(--line);background:#eef2ff;color:#323b6b;cursor:pointer}
.minibtn:hover{background:#e0e7ff}
table{width:100%;border-collapse:collapse}
th,td{border:1px solid var(--line);padding:8px;text-align:center}
th{background:#e9f5ff;color:#0f172a}
.small{font-size:12px;color:var(--sub)}
.modal{position:fixed;inset:0;background:rgba(0,0,0,.30);display:none;align-items:center;justify-content:center;z-index:9999}
.modal .box{background:#ffffff;border:1px solid var(--line);border-radius:14px;padding:16px;min-width:360px}
.badge{padding:3px 8px;border-radius:999px;font-size:12px}
.hi{background:#ffe4e6;color:#9f1239}
.md{background:#fff7ed;color:#9a3412}
.lo{background:#ecfdf5;color:#065f46}
"""

# ------------------------- UI Pages -------------------------
@app.get("/", response_class=HTMLResponse)
def root_redirect():
    return HTMLResponse('<meta http-equiv="refresh" content="0; url=/dashboard">')

@app.get("/ui", response_class=HTMLResponse)
def ui_page():
    html = """
<!doctype html><html lang="ar" dir="rtl"><head>
<meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>واجهة الإدخال — AI Maintenance Tracker</title>
<style>""" + BASE_STYLE + """</style></head><body>
<div class="wrap">
  <div class="row">
    <div class="col" style="flex:2 1 640px">
      <div class="card">
        <div class="row">
          <div class="col">
            <label>الموقع</label>
            <div class="inline-actions">
              <select id="selLocation" style="flex:1"></select>
              <button id="btnNewLoc" class="minibtn" title="إضافة موقع">➕</button>
            </div>
            <div class="inline-actions" style="margin-top:6px">
              <input id="txtSearch" placeholder="اكتب للبحث عن موقع (تحديث تلقائي)…"/>
              <button id="btnSearch" class="minibtn">بحث</button>
            </div>
          </div>
          <div class="col">
            <label>الفريق</label>
            <div class="inline-actions">
              <select id="selTeam" style="flex:1"></select>
              <button id="btnNewTeam" class="minibtn" title="إضافة فريق">➕</button>
            </div>
          </div>
        </div>

        <div class="row" style="margin-top:6px">
          <div class="col"><label>الموديل (المولد)</label><input id="genModel"/></div>
          <div class="col"><label>الرقم التسلسلي للمولد</label><input id="genSerial"/></div>
          <div class="col"><label>المصدر الثاني للطاقة</label>
            <select id="secondaryPower">
              <option>مولد + كهرباء عمومية</option>
              <option>مولد + طاقة شمسية</option>
            </select>
          </div>
        </div>

        <div class="row" style="margin-top:6px">
          <div class="col"><label>فاصل تبديل الزيت حسب الشركة (ساعات)</label><input id="oilIntervalHrs" type="number" min="1" value="250"/></div>
          <div class="col"><label>ساعات التشغيل المتوقعة يومياً</label><input id="dailyHours" type="number" min="1" value="6"/></div>
          <div class="col"><label>الموعد القادم (تقديري)</label><input id="nextDueDate" disabled/></div>
          <div class="col"><label>&nbsp;</label><button class="btn ok" id="btnSaveAsset">حفظ بيانات الأصل</button></div>
        </div>

        <div class="row" style="margin-top:6px">
          <div class="col"><a class="btn ghost" href="/dashboard" target="_blank">لوحة المتابعة</a></div>
          <div class="col"><a class="btn ghost" href="/calendar" target="_blank">تقويم أسبوعين</a></div>
          <div class="col"><a class="btn ghost" href="/predictions" target="_blank">خطة التنبؤ (الشهر القادم)</a></div>
        </div>
      </div>

      <div class="card">
        <div class="row"><div class="col"><strong>أعمال الصيانة</strong></div><div class="col"><span class="small" id="saveMsg"></span></div></div>
        <div class="row">
          <div class="col"><label>نوع الزيارة</label><select id="visitType">
            <option value="periodic">صيانة دورية</option><option value="emergency">صيانة طارئة</option>
            <option value="inspection">صيانة تفقدية</option><option value="supply">توريد مواد</option><option value="other">مهام أخرى</option>
          </select></div>
          <div class="col"><label>تاريخ الزيارة</label><input id="visitDate" type="date"/></div>
          <div class="col"><label>عداد المولد</label><input id="hourmeter" type="number" min="0"/></div>
          <div class="col"><label>عداد الكهرباء</label><input id="gridmeter" type="number" min="0"/></div>
        </div>
        <div class="row">
          <div class="col"><label>L1 (A)</label><input id="i1" type="number" step="0.01"/></div>
          <div class="col"><label>L2 (A)</label><input id="i2" type="number" step="0.01"/></div>
          <div class="col"><label>L3 (A)</label><input id="i3" type="number" step="0.01"/></div>
          <div class="col"><label>تغيير الزيت؟</label><select id="oil"><option value="false">لا</option><option value="true">نعم</option></select></div>
        </div>
        <div class="row">
          <div class="col"><label>الزيت (لتر)</label><input id="oilLiters" type="number" step="0.1" value="0"/></div>
          <div class="col"><label>فلتر الزيت</label><input id="oilFilter" type="number" value="0"/></div>
          <div class="col"><label>فلتر الديزل</label><input id="dieselFilter" type="number" value="0"/></div>
          <div class="col"><label>فلتر الهواء</label><input id="airFilter" type="number" value="0"/></div>
          <div class="col"><label>سير مولد</label><input id="beltQty" type="number" value="0"/></div>
        </div>
        <div class="row">
          <div class="col"><label>مصدر البلاغ</label><input id="emSource"/></div>
          <div class="col"><label>الإنذار</label><input id="emAlarm"/></div>
          <div class="col"><label>تصنيف المشكلة</label><input id="emClass" placeholder="ATS, Inverter, Fuel..."/></div>
        </div>
        <div class="row">
          <div class="col"><label>المنفّذ للعمل</label><input id="executor"/></div>
          <div class="col"><label>السائق</label><input id="driver"/></div>
          <div class="col"><label>تبعية الموقع</label><input id="owner"/></div>
        </div>
        <div class="row">
          <div class="col"><label>العمل المنجز</label><textarea id="summary"></textarea></div>
          <div class="col"><label>ملاحظات</label><textarea id="notes"></textarea></div>
        </div>
        <div class="row"><div class="col"><button class="btn" id="btnSaveVisit">حفظ الزيارة</button></div></div>
      </div>

      <div class="card">
        <div class="row"><div class="col"><strong>تصدير (Excel)</strong></div></div>
        <div class="row">
          <div class="col"><label>الشهر</label><input id="month" type="month"/></div>
          <div class="col"><label>&nbsp;</label><button class="btn" id="btnExport">تصدير الشهري</button></div>
          <div class="col"><label>&nbsp;</label><button class="btn secondary" id="btnLast30">آخر 30 يوماً</button></div>
        </div>
      </div>
    </div>

    <div class="col" style="flex:1 1 380px">
      <div class="card">
        <div class="row"><div class="col"><strong>معاينة Excel</strong></div></div>
        <table id="preview"><thead><tr><th>العمود</th><th>القيمة</th></tr></thead><tbody></tbody></table>
      </div>
      <div class="card">
        <div class="row"><div class="col"><strong>الإنذارات المفتوحة</strong></div><div class="col" style="text-align:left"><button class="btn secondary" id="btnReloadAlerts">تحديث</button></div></div>
        <div id="alertsBox" class="small">لا توجد إنذارات.</div>
      </div>
    </div>
  </div>
</div>

<!-- Modal: Location -->
<div id="modalLoc" class="modal">
  <div class="box">
    <div class="row"><div class="col"><strong>إضافة موقع</strong></div><div class="col" style="text-align:left"><button class="btn ghost" onclick="closeModal('modalLoc')">إغلاق</button></div></div>
    <div class="row"><div class="col"><label>الاسم</label><input id="newLocName"/></div><div class="col"><label>المنطقة</label><input id="newLocRegion"/></div><div class="col"><label>تبعية</label><input id="newLocOwner"/></div></div>
    <div class="row"><div class="col"><button class="btn ok" id="doCreateLoc">حفظ</button></div></div>
  </div>
</div>

<!-- Modal: Team -->
<div id="modalTeam" class="modal">
  <div class="box">
    <div class="row"><div class="col"><strong>إضافة فريق</strong></div><div class="col" style="text-align:left"><button class="btn ghost" onclick="closeModal('modalTeam')">إغلاق</button></div></div>
    <div class="row"><div class="col"><label>اسم الفريق</label><input id="newTeamName"/></div></div>
    <div class="row"><div class="col"><button class="btn ok" id="doCreateTeam">حفظ</button></div></div>
  </div>
</div>

<script>
function el(id){return document.getElementById(id);}
function show(id){el(id).style.display="flex";}
function hide(id){el(id).style.display="none";}
function modal(id){show(id);} function closeModal(id){hide(id);}
async function api(path,opts={}){const h=opts.headers||{};h["Content-Type"]=h["Content-Type"]||"application/json";return fetch(path,{...opts,headers:h});}
function debounce(fn,ms){let t;return(...a)=>{clearTimeout(t);t=setTimeout(()=>fn.apply(this,a),ms);}}

/* load data */
async function loadLocations(q=""){const locs=await fetch("/locations"+(q?`?q=${encodeURIComponent(q)}`:"")).then(r=>r.json()); el("selLocation").innerHTML=locs.map(l=>`<option value="${l.id}" data-region="${l.region}">${l.region} — ${l.name}</option>`).join(""); if(locs[0]) loadAsset(locs[0].id);}
async function loadTeams(){const t=await fetch("/teams").then(r=>r.json()); el("selTeam").innerHTML=t.map(x=>`<option value="${x.id}">${x.name}</option>`).join("");}
async function loadAsset(locationId){
  const a = await fetch(`/assets/by_location?location_id=${locationId}`).then(r=>r.json());
  if(!a) return;
  el("genModel").value = a.generator_model || "";
  el("genSerial").value = a.generator_serial || "";
  el("secondaryPower").value = a.secondary_power || "مولد + كهرباء عمومية";
  el("oilIntervalHrs").value = a.oil_interval_hours || 250;
  el("dailyHours").value = a.daily_operating_hours || 6;
  el("nextDueDate").value = a.next_oil_due_date ? new Date(a.next_oil_due_date).toLocaleDateString() : "";

  el("btnSaveAsset").onclick = async () => {
    const body = {
      generator_model: el("genModel").value,
      generator_serial: el("genSerial").value,
      secondary_power: el("secondaryPower").value,
      oil_interval_hours: parseInt(el("oilIntervalHrs").value||"250",10),
      daily_operating_hours: parseInt(el("dailyHours").value||"6",10)
    };
    const res = await api(`/assets/${a.id}`, { method:"PUT", body: JSON.stringify(body) });
    if(res.ok){
      alert("تم حفظ بيانات الأصل");
      const refreshed = await fetch(`/assets/by_location?location_id=${locationId}`).then(r=>r.json());
      el("nextDueDate").value = refreshed.next_oil_due_date ? new Date(refreshed.next_oil_due_date).toLocaleDateString() : "";
    }else{
      let msg="تعذر حفظ الأصل"; try{const j=await res.json(); if(j.detail){msg+=" — "+(typeof j.detail==="string"?j.detail:JSON.stringify(j.detail));}}catch(e){}; alert(msg);
    }
  };
}

el("selLocation").addEventListener("change",e=>loadAsset(parseInt(e.target.value)));
el("btnSearch").onclick=()=>loadLocations(el("txtSearch").value.trim());
el("txtSearch").addEventListener("keyup", debounce(()=>loadLocations(el("txtSearch").value.trim()), 250));

/* modals create */
el("btnNewLoc").onclick=()=>modal("modalLoc");
el("btnNewTeam").onclick=()=>modal("modalTeam");
el("doCreateLoc").onclick = async () => {
  const name = (el("newLocName").value || "").trim();
  const region = (el("newLocRegion").value || "").trim();
  const owner = (el("newLocOwner").value || "").trim();

  if (!name || !region) {
    alert("الرجاء إدخال اسم الموقع والمنطقة");
    return;
  }

  const body = { name, region, site_owner: owner || null };
  const res = await api("/locations", { method: "POST", body: JSON.stringify(body) });

  if (res.ok) {
    closeModal("modalLoc");
    await loadLocations(el("txtSearch").value.trim());
  } else {
    let msg = "تعذر إضافة الموقع";
    try {
      const err = await res.json();
      if (err && err.detail) msg = err.detail;     // نعرض سبب الخطأ القادم من الخادم
    } catch (_) {
      try { msg = await res.text(); } catch (_) {}
    }
    alert(msg);
  }
};

 try{const j=await res.json(); if(j.detail){msg+=" — "+(typeof j.detail==="string"?j.detail:JSON.stringify(j.detail));}}catch(e){}; alert(msg); }
};
el("doCreateTeam").onclick=async()=>{const body={name:el("newTeamName").value}; const res=await api("/teams",{method:"POST",body:JSON.stringify(body)}); if(res.ok){closeModal("modalTeam"); loadTeams();} else alert("تعذر إضافة الفريق");};

/* save visit */
el("btnSaveVisit").onclick=async()=>{
  el("saveMsg").textContent="جارٍ الحفظ...";
  const body={
    location_id: parseInt(el("selLocation").value,10), team_id: parseInt(el("selTeam").value,10),
    visit_type: el("visitType").value,
    date: el("visitDate").value ? (el("visitDate").value + "T08:00:00") : undefined,
    hourmeter: parseInt(el("hourmeter").value||"0",10), grid_meter: parseInt(el("gridmeter").value||"0",10),
    i_l1: parseFloat(el("i1").value||"0"), i_l2: parseFloat(el("i2").value||"0"), i_l3: parseFloat(el("i3").value||"0"),
    performed_oil_change: el("oil").value==="true",
    oil_liters: parseFloat(el("oilLiters").value||"0"), oil_filter: parseInt(el("oilFilter").value||"0",10),
    diesel_filter: parseInt(el("dieselFilter").value||"0",10), air_filter: parseInt(el("airFilter").value||"0",10), belt_qty: parseInt(el("beltQty").value||"0",10),
    summary: el("summary").value, executor_name: el("executor").value, driver_name: el("driver").value,
    site_owner: el("owner").value, emergency_source: el("emSource").value, emergency_alarm: el("emAlarm").value,
    emergency_class: el("emClass").value, notes: el("notes").value
  };
  const res=await api("/visits",{method:"POST",body:JSON.stringify(body)}); let out={}; try{out=await res.json();}catch(e){}
  el("saveMsg").textContent=res.ok?("تم الحفظ (Visit# "+(out.visit_id||"")+")"):"خطأ في الحفظ";
  if(res.ok){
    el("btnReloadAlerts").click();
    const locId=parseInt(el("selLocation").value,10); await loadAsset(locId); // refresh next due
  }
  preview();
};

/* export */
el("btnExport").onclick=()=>{const m=el("month").value; if(!m){alert("اختر الشهر");return;} window.location.href="/reports/monthly?month="+encodeURIComponent(m);};
el("btnLast30").onclick=()=>{const end=new Date(); const start=new Date(); start.setDate(start.getDate()-30); const fmt=d=>d.toISOString().slice(0,10); window.location.href=`/reports/export?start=${fmt(start)}&end=${fmt(end)}`;};

/* alerts */
el("btnReloadAlerts").onclick=async()=>{const a=await fetch("/alerts").then(r=>r.json()); if(!a.length){el("alertsBox").textContent="لا توجد إنذارات.";return;} const rows=a.map(x=>`<tr><td>${new Date(x.created_at).toLocaleString()}</td><td>${x.location_id}</td><td>${x.kind}</td><td>${x.message}</td><td>${x.level}</td></tr>`).join(""); el("alertsBox").innerHTML=`<table><thead><tr><th>التاريخ</th><th>الموقع</th><th>النوع</th><th>الرسالة</th><th>المستوى</th></tr></thead><tbody>${rows}</tbody></table>`;};

/* preview */
function preview(){
  const opt=el("selLocation").selectedOptions[0]; const region=opt?opt.dataset.region:""; const locTxt=opt?opt.textContent:"";
  const vt=el("visitType").value; const type=vt==="emergency"?"صيانة طارئة":vt==="periodic"?"صيانة دورية":vt==="inspection"?"صيانة تفقدية":vt==="supply"?"توريد مواد":"مهام أخرى";
  const map=[["التاريخ", el("visitDate").value || new Date().toISOString().slice(0,10)],["المنطقة",region],["الموقع",locTxt],["نوع العمل",type],["العمل المنجز (ملخص فقط)",el("summary").value],["عداد الساعات",el("hourmeter").value||0],["فارق القراءه","(يُحتسب تلقائياً)"],["الزيت(لتر)",el("oilLiters").value||0],["فلتر الزيت",el("oilFilter").value||0],["فلتر الديزل",el("dieselFilter").value||0],["فلتر الهواء",el("airFilter").value||0],["سير مولد",el("beltQty").value||0],["L1",el("i1").value||0],["L2",el("i2").value||0],["L3",el("i3").value||0],["اسم القطعه",""],["الكميه",""],["الإنذار",el("emAlarm").value||""],["مصدر البلاغ",el("emSource").value||""],["تصنيف المشكلة",el("emClass").value||""],["تبعية الموقع",el("owner").value||""],["المنفذ للعمل",el("executor").value||""],["السائق",el("driver").value||""],["ملاحظات",el("notes").value||""]];
  document.querySelector("#preview tbody").innerHTML=map.map(([k,v])=>`<tr><td>${k}</td><td>${v}</td></tr>`).join("");
}
["selLocation","visitType","visitDate","hourmeter","gridmeter","i1","i2","i3","oil","oilLiters","oilFilter","dieselFilter","airFilter","beltQty","emSource","emAlarm","emClass","executor","driver","owner","summary","notes"].forEach(id=>{const e=el(id); if(e) e.addEventListener("input", preview);});

/* boot */
(async function(){ await loadLocations(); await loadTeams(); preview(); })();
</script>
</body></html>
"""
    return HTMLResponse(html)

PRED_STYLE = BASE_STYLE + ".tag{padding:3px 8px;border-radius:999px;background:#e4f1ff;color:#0b2b4b}.due{background:#fef3c7;color:#92400e}.late{background:#fee2e2;color:#991b1b}"
@app.get("/predictions", response_class=HTMLResponse)
def predictions_page():
    html = """
<!doctype html><html lang="ar" dir="rtl"><head>
<meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>خطة التنبؤ للشهر القادم</title>
<style>""" + PRED_STYLE + """</style></head><body>
<div class="wrap">
  <h2>خانة التنبؤ (الشهر القادم)</h2>
  <div class="row">
    <div class="col"><a class="btn ghost" href="/ui">العودة إلى إدخال البيانات</a></div>
    <div class="col"><a class="btn secondary" href="/predictions/export">تنزيل خطة التنبؤ (Excel)</a></div>
  </div>
  <div id="box" class="card">جارٍ التحميل...</div>
</div>
<script>
(async function(){
  const data = await fetch('/predictions/next_month').then(r=>r.json());
  if(!data.length){ document.getElementById('box').innerHTML='لا توجد بيانات.'; return; }
  const rows = data.map(x=>{
    const due = x.next_due_date ? new Date(x.next_due_date).toLocaleDateString() : '-';
    const days = (x.days_left===null||x.days_left===undefined)?'-':x.days_left;
    let chip = '<span class="tag">خارج الإطار</span>';
    if(x.days_left!==null){
      if(x.days_left < 0) chip = '<span class="tag late">متجاوز</span>';
      else if(x.days_left <= 30) chip = '<span class="tag due">خلال 30 يوماً</span>';
    }
    return `<tr>
      <td>${x.region}</td><td>${x.location}</td>
      <td>${x.hourmeter}</td><td>${x.interval_hours}</td><td>${x.daily_hours}</td>
      <td>${x.last_oil_hours||'-'}</td><td>${x.last_oil_date? new Date(x.last_oil_date).toLocaleDateString() : '-'}</td>
      <td>${due}</td><td>${days}</td><td>${chip}</td>`;
  }).join('');
  document.getElementById('box').innerHTML = `
    <table>
      <thead>
        <tr>
          <th>المنطقة</th><th>الموقع</th>
          <th>عداد الساعات</th><th>فاصل الشركة (ساعات)</th><th>التشغيل اليومي (س)</th>
          <th>آخر تبديل زيت (س)</th><th>آخر تبديل زيت (ت)</th>
          <th>الموعد القادم</th><th>أيام متبقية</th><th>الحالة</th>
        </tr>
      </thead>
      <tbody>${rows}</tbody>
    </table>`;
})();
</script>
</body></html>
"""
    return HTMLResponse(html)

# ---------------------- Dashboard / Calendar ----------------------
DASH_STYLE = BASE_STYLE + ".kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}.k{background:#ffffff;border:1px solid var(--line);border-radius:16px;padding:14px;box-shadow:0 6px 18px rgba(0,0,0,.05)}.k .v{font-size:26px;font-weight:800;color:#0f172a}.k .t{color:#64748b;font-size:12px}"
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard_page():
    html = """
<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>لوحة المتابعة</title><style>""" + DASH_STYLE + """</style></head><body>
<div class="wrap">
  <h2>لوحة متابعة المواقع</h2>
  <div id="kpis" class="kpis"></div>
  <div class="row">
    <div class="col"><div class="card"><div class="row"><div class="col"><strong>خطة اليوم</strong></div><div class="col"><select id="region"></select><button class="btn" id="btnReloadPlan">تحديث</button></div></div><div id="planBox"></div></div></div>
    <div class="col"><div class="card"><div class="row"><div class="col"><strong>صحة المواقع</strong></div><div class="col"><button class="btn secondary" id="btnReloadHealth">تحديث</button></div></div><div id="healthBox"></div></div></div>
  </div>
</div>
<script>
const $=id=>document.getElementById(id);
const lvl=s=> s>=80?'<span class="badge hi">حرِج</span>':(s>=60?'<span class="badge md">مرتفع</span>':'<span class="badge lo">منخفض</span>');
async function loadKPIs(){const k=await fetch('/kpi/overview').then(r=>r.json()); $('kpis').innerHTML=`<div class="k"><div class="v">${k.locations}</div><div class="t">إجمالي المواقع</div></div><div class="k"><div class="v">${k.alerts_open}</div><div class="t">إنذارات مفتوحة</div></div><div class="k"><div class="v">${k.emergency_30d}</div><div class="t">بلاغات طارئة (30ي)</div></div><div class="k"><div class="v">${k.overdue_periodic}</div><div class="t">صيانة دورية متأخرة</div></div>`;}
async function loadRegions(){const locs=await fetch('/locations').then(r=>r.json()); const regs=[...new Set(locs.map(l=>l.region))]; $('region').innerHTML='<option value="">كل المناطق</option>'+regs.map(r=>`<option>${r}</option>`).join('');}
async function loadPlan(){const reg=$('region').value; const url=reg?`/planner/today?region=${encodeURIComponent(reg)}`:'/planner/today'; const list=await fetch(url).then(r=>r.json()); if(!list.length){$('planBox').innerHTML='لا توجد زيارات مقترحة اليوم.';return;} const rows=list.map(i=>`<tr><td>${i.region}</td><td>${i.location}</td><td>${i.last_visit_at?new Date(i.last_visit_at).toLocaleDateString():'-'}</td><td>${i.recent_emergencies_30d}</td><td>${i.due_oil?'نعم':'-'}</td><td>${i.has_unbalance?'نعم':'-'}</td><td>${i.reasons.join('، ')}</td><td>${i.score} ${lvl(i.score)}</td></tr>`).join(''); $('planBox').innerHTML=`<table><thead><tr><th>المنطقة</th><th>الموقع</th><th>آخر زيارة</th><th>طارئ (30ي)</th><th>زيت</th><th>عدم توازن</th><th>الأسباب</th><th>الأولوية</th></tr></thead><tbody>${rows}</tbody></table>`;}
async function loadHealth(){const list=await fetch('/health/locations').then(r=>r.json()); list.sort((a,b)=>b.score-a.score); const rows=list.map(i=>`<tr><td>${i.region}</td><td>${i.location}</td><td>${i.last_visit_at?new Date(i.last_visit_at).toLocaleDateString():'-'}</td><td>${i.recent_emergencies_30d}</td><td>${i.due_oil?'نعم':'-'}</td><td>${i.has_unbalance?'نعم':'-'}</td><td>${i.recommend_in_days==0?'اليوم':(i.recommend_in_days+' يوم')}</td><td>${i.score} ${lvl(i.score)}</td></tr>`).join(''); $('healthBox').innerHTML=`<table><thead><tr><th>المنطقة</th><th>الموقع</th><th>آخر زيارة</th><th>طارئ (30ي)</th><th>زيت</th><th>عدم توازن</th><th>موعد مقترح</th><th>المخاطر</th></tr></thead><tbody>${rows}</tbody></table>`;}
$('btnReloadPlan').onclick=loadPlan; $('btnReloadHealth').onclick=loadHealth; (async function(){await loadKPIs();await loadRegions();await loadPlan();await loadHealth();})();</script>
</body></html>
"""
    return HTMLResponse(html)

CAL_STYLE = BASE_STYLE + ".daygrid{display:grid;grid-template-columns:repeat(7,1fr);gap:8px}.day{background:#ffffff;border:1px solid var(--line);border-radius:12px;padding:8px;min-height:140px;box-shadow:0 4px 14px rgba(0,0,0,.04)}.badge{padding:3px 6px;border-radius:999px;background:#e0e7ff;color:#1e293b}"
@app.get("/calendar", response_class=HTMLResponse)
def calendar_page():
    html = """
<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>تقويم أسبوعين</title><style>""" + CAL_STYLE + """</style></head><body>
<div class="wrap"><h2>خطة أسبوعين (موزعة على الفرق)</h2><div id="grid"></div></div>
<script>
(async function(){
  const data=await fetch('/planner/fortnight').then(r=>r.json());
  const teams=data.teams; const days=data.days.map(d=>new Date(d).toLocaleDateString());
  let html='<div class="daygrid">';
  for(let i=0;i<days.length;i++){
    const key=new Date(data.days[i]).toISOString().slice(0,10);
    html+=`<div class="day"><div><strong>${days[i]}</strong> <span class="badge">سعة الفريق: ${data.team_capacity}/يوم</span></div>`;
    for(const t of teams){
      const items=data.schedule[key][t.id]||[];
      if(items.length){
        html+=`<div style="margin-top:6px"><span class="badge">${t.name}</span><ul style="margin:6px 0 0 12px">`+items.map(x=>`<li>${x.region} — ${x.location} (أولوية ${x.score})</li>`).join("")+`</ul></div>`;
      }
    }
    html+='</div>';
  }
  html+='</div>';
  document.getElementById('grid').innerHTML=html;
})();
</script>
</body></html>
"""
    return HTMLResponse(html)

# ---------------------- PWA (optional) -------------------
@app.get("/manifest.json")
def manifest():
    return JSONResponse({"name":APP_NAME,"short_name":"AI Tracker","start_url":"/dashboard","display":"standalone",
                         "background_color":"#f6f8fc","theme_color":"#4f46e5","icons":[]})

@app.get("/sw.js")
def sw():
    code = "self.addEventListener('install',e=>{e.waitUntil(caches.open('ai-tracker-v1').then(c=>c.addAll(['/','/ui','/dashboard','/calendar','/predictions'])))});self.addEventListener('fetch',e=>{e.respondWith(caches.match(e.request).then(r=>r||fetch(e.request)))})"
    return PlainTextResponse(code, media_type="application/javascript")

# ---------------------- Scheduler ------------------------
def run_daily_due_checks():
    with Session(engine) as s:
        for a in s.exec(select(Asset)).all():
            if due_oil(a):
                existing = s.exec(select(Alert).where(Alert.location_id==a.location_id)
                                  .where(Alert.kind=="due_oil").where(Alert.is_open==True)).first()
                if not existing:
                    s.add(Alert(location_id=a.location_id, kind="due_oil", level="info", message="زيت/فلاتر مستحقة"))
            a.next_oil_due_date = compute_next_oil_due(a)  # refresh predicted due daily
            s.add(a)
        s.commit()

def start_scheduler():
    sch=BackgroundScheduler()
    sch.add_job(run_daily_due_checks, CronTrigger(hour=8, minute=0))
    sch.start()

# ---------------------- Startup --------------------------
@app.on_event("startup")
def on_startup():
    SQLModel.metadata.create_all(engine)  # ensure tables exist
    migrate_db()                          # add missing columns (safe)
    init_db()                             # seed if empty
    start_scheduler()
